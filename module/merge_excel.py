# module/merge_excel.py
import os
import openpyxl
import copy
import logging
import re

logger = logging.getLogger(__name__)

EVENT_NAME_MAP = {
    "50fr.xlsx": "50m 自由形",
    "100fr.xlsx": "100m 自由形",
    "200fr.xlsx": "200m 自由形",
    "400fr.xlsx": "400m 自由形",
    "50ba.xlsx": "50m 背泳ぎ",
    "100ba.xlsx": "100m 背泳ぎ",
    "200ba.xlsx": "200m 背泳ぎ",
    "50br.xlsx": "50m 平泳ぎ",
    "100br.xlsx": "100m 平泳ぎ",
    "200br.xlsx": "200m 平泳ぎ",
    "50fly.xlsx": "50m バタフライ",
    "100fly.xlsx": "100m バタフライ",
    "200fly.xlsx": "200m バタフライ",
    "200im.xlsx": "200m 個人メドレー",
    "400im.xlsx": "400m 個人メドレー",
}

def offset_formula(formula_str, row_offset):
    if not formula_str or not isinstance(formula_str, str) or not formula_str.startswith("="):
        return formula_str
    
    def replace_match_safe(m):
        col = m.group(1)
        row_str = m.group(2)
        
        # Check if there is $ before column
        col_start = m.start(1)
        if col_start > 0 and formula_str[col_start - 1] == '$':
            return m.group(0)
            
        # Check if there is $ before row
        row_start = m.start(2)
        if row_start > 0 and formula_str[row_start - 1] == '$':
            return m.group(0)
            
        # Check if followed by (
        end_idx = m.end()
        if end_idx < len(formula_str) and formula_str[end_idx] == '(':
            return m.group(0)
            
        row = int(row_str)
        return f"{col}{row + row_offset}"
        
    pattern = r'\b([A-Z]+)(\d+)\b'
    return re.sub(pattern, replace_match_safe, formula_str)

def copy_cell(src_cell, dest_cell, event_name, row_offset):
    """セル値とスタイルをコピーする（数式のオフセットと種目名置換を含む）"""
    val = src_cell.value
    if isinstance(val, str) and val.startswith("="):
        val_clean = val.replace(" ", "").replace("$", "")
        if val_clean == "=A2":
            dest_cell.value = event_name
        else:
            dest_cell.value = offset_formula(val, row_offset)
    else:
        dest_cell.value = val

    if src_cell.has_style:
        dest_cell.font = copy.copy(src_cell.font)
        dest_cell.border = copy.copy(src_cell.border)
        dest_cell.fill = copy.copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.alignment = copy.copy(src_cell.alignment)

def merge_excel_files(result_folder: str, file_order: list[str], output_filename: str = "combined_program.xlsx") -> str:
    """
    指定された順序で複数の種目別Excelファイルを縦に結合し、一つのファイルに出力する。
    
    引数:
        - result_folder: 種目別Excelファイルが格納されているフォルダパス
        - file_order: 結合するファイル名（例：["50fr.xlsx", "100fr.xlsx"]）のリスト（順番指定）
        - output_filename: 出力する結合ファイル名
        
    戻り値:
        - 作成された結合ファイルのフルパス
    """
    try:
        new_wb = openpyxl.Workbook()
        dest_ws = new_wb.active
        dest_ws.title = "一括印刷用プログラム"
        
        # テンプレートファイル template.xlsx からカバーシート（大会要領）をコピーして先頭に挿入する
        template_path = "template.xlsx"
        if os.path.exists(template_path):
            try:
                temp_wb = openpyxl.load_workbook(template_path)
                cover_sheet_name = temp_wb.sheetnames[0]
                # 競技シート以外であればカバーシートと判定してコピー
                if cover_sheet_name not in ["50m", "100m", "200m", "400m"]:
                    logger.info(f"カバーシート '{cover_sheet_name}' をコピーします")
                    cover_src_ws = temp_wb[cover_sheet_name]
                    cover_dest_ws = new_wb.create_sheet(title=cover_sheet_name, index=0)
                    
                    # 内容とスタイルをコピー
                    for r in range(1, cover_src_ws.max_row + 1):
                        h = cover_src_ws.row_dimensions[r].height
                        if h:
                            cover_dest_ws.row_dimensions[r].height = h
                        for c in range(1, cover_src_ws.max_column + 1):
                            src_cell = cover_src_ws.cell(row=r, column=c)
                            dest_cell = cover_dest_ws.cell(row=r, column=c)
                            dest_cell.value = src_cell.value
                            if src_cell.has_style:
                                dest_cell.font = copy.copy(src_cell.font)
                                dest_cell.border = copy.copy(src_cell.border)
                                dest_cell.fill = copy.copy(src_cell.fill)
                                dest_cell.number_format = src_cell.number_format
                                dest_cell.alignment = copy.copy(src_cell.alignment)
                                
                    # 結合セルのコピー
                    for merged_range in cover_src_ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        cover_dest_ws.merge_cells(
                            start_row=min_row,
                            start_column=min_col,
                            end_row=max_row,
                            end_column=max_col
                        )
            except Exception as e:
                logger.error(f"カバーシートのコピー中にエラー: {e}", exc_info=True)
                
        current_row = 1
        
        # カラム幅の最大値を追跡して設定する
        max_col_widths = {}
        
        for idx, filename in enumerate(file_order):
            if filename.startswith("~$"):
                continue
            file_path = os.path.join(result_folder, filename)
            if not os.path.exists(file_path):
                logger.warning(f"結合対象ファイルが見つかりません: {file_path}")
                continue
                
            logger.info(f"結合中: {file_path} (開始行: {current_row})")
            # data_only=False で読み込んで数式を保持する
            src_wb = openpyxl.load_workbook(file_path, data_only=False)
            src_ws = src_wb.active
            
            # 各ファイルの列幅を収集
            for col_idx in range(1, src_ws.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                width = src_ws.column_dimensions[col_letter].width
                if width:
                    max_col_widths[col_letter] = max(max_col_widths.get(col_letter, 0), width)
            
            # 種目名とイベント番号の文字列をフォーマット
            japanese_event_name = EVENT_NAME_MAP.get(filename.lower(), filename)
            event_name = f"No. {idx + 1} {japanese_event_name}"
            
            # 2回目以降の結合時の改ページ判定・区切り行設定
            if idx > 0:
                prev_filename = file_order[idx - 1]
                is_prev_long = prev_filename.startswith("200") or prev_filename.startswith("400")
                is_curr_long = filename.startswith("200") or filename.startswith("400")
                
                if is_prev_long and is_curr_long:
                    # 200m以上の種目が連続する場合は改ページして次のページから開始
                    from openpyxl.worksheet.pagebreak import Break
                    dest_ws.row_breaks.append(Break(id=current_row - 1))
                else:
                    # それ以外は詰め込み用に2行のスペースを挿入
                    dest_ws.row_dimensions[current_row].height = 20
                    current_row += 2
                
            row_offset = current_row - 1
            
            # 改ページ情報をコピー
            from openpyxl.worksheet.pagebreak import Break
            for brk in src_ws.row_breaks.brk:
                new_break_row = brk.id + row_offset
                dest_ws.row_breaks.append(Break(id=new_break_row))
            
            # 結合セルの情報をコピー
            for merged_range in src_ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                new_min_row = min_row + row_offset
                new_max_row = max_row + row_offset
                dest_ws.merge_cells(
                    start_row=new_min_row,
                    start_column=min_col,
                    end_row=new_max_row,
                    end_column=max_col
                )
            
            # 行データとスタイルをコピー
            for r in range(1, src_ws.max_row + 1):
                # 行の高さをコピー
                src_height = src_ws.row_dimensions[r].height
                if src_height:
                    dest_ws.row_dimensions[r + row_offset].height = src_height
                    
                for c in range(1, src_ws.max_column + 1):
                    src_cell = src_ws.cell(row=r, column=c)
                    dest_cell = dest_ws.cell(row=r + row_offset, column=c)
                    copy_cell(src_cell, dest_cell, event_name, row_offset)
                    
            current_row += src_ws.max_row
            
        # 左右の列幅を対称にする（BとG, CとH, DとI, EとJ）
        wB = max_col_widths.get('B', 0)
        wG = max_col_widths.get('G', 0)
        if wB or wG:
            max_w = max(wB, wG)
            max_col_widths['B'] = max_w
            max_col_widths['G'] = max_w

        wC = max_col_widths.get('C', 0)
        wH = max_col_widths.get('H', 0)
        if wC or wH:
            max_w = max(wC, wH)
            max_col_widths['C'] = max_w
            max_col_widths['H'] = max_w

        wD = max_col_widths.get('D', 0)
        wI = max_col_widths.get('I', 0)
        if wD or wI:
            max_w = max(wD, wI)
            max_col_widths['D'] = max_w
            max_col_widths['I'] = max_w

        wE = max_col_widths.get('E', 0)
        wJ = max_col_widths.get('J', 0)
        if wE or wJ:
            max_w = max(wE, wJ, 8.43)
            max_col_widths['E'] = max_w
            max_col_widths['J'] = max_w

        # 最終的な列幅を設定
        for col_letter, width in max_col_widths.items():
            dest_ws.column_dimensions[col_letter].width = width
            
        # 印刷範囲をA1から最終セルまでに設定
        max_col_letter = openpyxl.utils.get_column_letter(dest_ws.max_column)
        dest_ws.print_area = f"A1:{max_col_letter}{dest_ws.max_row}"
            
        # 印刷設定のコピー（余白などの基本設定を適用）
        dest_ws.page_setup.orientation = dest_ws.ORIENTATION_PORTRAIT # 縦向き印刷をデフォルトにする
        dest_ws.page_setup.paperSize = dest_ws.PAPERSIZE_A4
        
        # 最も狭い余白（左0.31, 右0.27, 上0.59, 下0.27インチ）を設定して印刷見切れを防ぐ
        dest_ws.page_margins.left = 0.31
        dest_ws.page_margins.right = 0.27
        dest_ws.page_margins.top = 0.59
        dest_ws.page_margins.bottom = 0.27
        dest_ws.page_margins.header = 0.31
        dest_ws.page_margins.footer = 0.31
        
        # 6組（60行）が縦に1ページに収まるように、また横幅も見切れないように印刷倍率を65%に設定
        dest_ws.page_setup.scale = 65
        dest_ws.sheet_properties.pageSetUpPr.fitToPage = False
        dest_ws.page_setup.fitToWidth = None
        dest_ws.page_setup.fitToHeight = None
        
        # ファイルを保存
        output_path = os.path.join(result_folder, output_filename)
        new_wb.save(output_path)
        logger.info(f"結合ファイルを保存しました: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"merge_excel_filesでエラーが発生しました: {e}", exc_info=True)
        raise
