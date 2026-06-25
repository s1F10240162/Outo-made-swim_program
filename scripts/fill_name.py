# fill_name.py
# 役割: Excelの選手IDを基に、選手の名前・フリガナ・学校名・学年を補完する。

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)
import logging
import traceback
from module.send_message import send_slack_message


def get_player_data_by_id(player_ids, csv_path):
    """
    IDリストに基づき、選手の名前・フリガナ・学校名・学年を取得する。

    引数:
        - player_ids: 選手IDのリスト
        - csv_path: 選手データが格納されたCSVファイルのパス

    戻り値:
        - 選手IDをキーとした辞書 {ID: (氏名, ﾌﾘｶﾞﾅ, 学校名, 学年)}
    """
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSVファイルが見つかりません: {csv_path}")

    df = pd.read_csv(csv_path)

    # ID列の存在確認
    if "ID" not in df.columns:
        raise ValueError(
            f"CSV に 'ID' 列が存在しません。利用可能な列: {df.columns.tolist()}"
        )

    # 必要な列の存在確認
    required_columns = ["ID", "氏名", "ﾌﾘｶﾞﾅ", "学校名", "学年"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"CSVに必要な列が不足しています: {missing_columns}")

    # ID列を文字列に変換
    df["ID"] = df["ID"].astype(str).str.strip()
    player_ids = [str(id_).strip() for id_ in player_ids if id_ is not None]

    # IDと選手データのマッピングを作成
    id_to_data = {}
    for _, row in df.iterrows():
        id_to_data[row["ID"]] = (
            row["氏名"] if pd.notna(row["氏名"]) else None,
            row["ﾌﾘｶﾞﾅ"] if pd.notna(row["ﾌﾘｶﾞﾅ"]) else None,
            row["学校名"] if pd.notna(row["学校名"]) else None,
            row["学年"] if pd.notna(row["学年"]) else None,
        )

    return {id_: id_to_data.get(id_, (None, None, None, None)) for id_ in player_ids}


def update_excel_with_player_data(excel_path, csv_path, output_dir):
    """
    Excelファイル内のすべての選手IDセルを自動検出して、名前・フリガナ・学校名・学年に置き換える。

    引数:
        - excel_path: 更新対象のExcelファイルパス
        - csv_path: 選手データのCSVファイルパス
        - output_dir: 更新後のExcelファイルを保存するディレクトリ
    """
    try:
        if not os.path.exists(excel_path):
            logging.error(f"Excelファイルが存在しません: {excel_path}")
            send_slack_message(
                os.getenv("APP_NAME", "AquaProgrammer"),
                f"Excelファイルが存在しません: {excel_path}",
            )
            return

        # CSVから選手データを取得
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"CSVファイルが見つかりません: {csv_path}")

        df = pd.read_csv(csv_path)
        df["ID"] = df["ID"].astype(str).str.strip()

        # IDと選手データのマッピングを作成
        id_to_data = {}
        for _, row in df.iterrows():
            id_to_data[row["ID"]] = (
                row["氏名"] if pd.notna(row["氏名"]) else None,
                row["ﾌﾘｶﾞﾅ"] if pd.notna(row["ﾌﾘｶﾞﾅ"]) else None,
                row["学校名"] if pd.notna(row["学校名"]) else None,
                row["学年"] if pd.notna(row["学年"]) else None,
            )

        # Excelファイルを読み込む
        wb = load_workbook(excel_path)
        ws = wb.active

        # Determine target ID columns based on filename
        filename = os.path.basename(excel_path)
        if filename.startswith("50") or filename.startswith("100"):
            id_cols = [2, 7]  # Column B, G
        else:
            id_cols = [2]      # Column B

        # 指定されたID列のみをスキャンして置換
        updated_count = 0
        for r in range(1, ws.max_row + 1):
            for c in id_cols:
                cell_value = ws.cell(row=r, column=c).value
                if cell_value is None:
                    continue

                cell_value_str = str(cell_value).strip()
                # 整数で、かつIDマップに存在する場合のみ置換
                if cell_value_str.isdigit() and cell_value_str in id_to_data:
                    name, hurigana, school, grade = id_to_data[cell_value_str]

                    # 氏名
                    ws.cell(row=r, column=c).value = name
                    # フリガナ (1列右)
                    if hurigana is not None:
                        ws.cell(row=r, column=c + 1).value = hurigana
                    # 学校名 (2列右)
                    if school is not None:
                        ws.cell(row=r, column=c + 2).value = school

                    updated_count += 1

        # 出力ディレクトリを作成
        os.makedirs(output_dir, exist_ok=True)

        # 保存
        original_filename = os.path.basename(excel_path)
        new_filename = original_filename.replace("_id.xlsx", ".xlsx")
        output_path = os.path.join(output_dir, new_filename)
        wb.save(output_path)
        logging.info(
            f"Excelファイルを更新・保存しました: {output_path} (更新数: {updated_count})"
        )

    except Exception as e:
        logging.error(f"Excel更新処理中にエラー: {e}", exc_info=True)
        send_slack_message(
            os.getenv("APP_NAME", "AquaProgrammer"),
            f"Excel更新処理中にエラー: {e}\n{traceback.format_exc()}",
        )
        raise


def main(result_output_folder=None, merged_csv_data_file=None):
    """
    メイン処理:
    1. CSVファイルから選手情報を取得
    2. 競技別にExcelシートを更新
    3. 更新完了のメッセージを出力
    """
    from dotenv import load_dotenv

    load_dotenv()

    # 環境変数から設定を読み込む
    result_output_folder = result_output_folder or os.getenv("RESULT_DATA_FILE", "result")
    directory_path = os.getenv("DIRECTORY_PATH", "test/")
    merged_csv_data_file = merged_csv_data_file or os.path.join(
        directory_path, os.getenv("MERGED_CSV_DATA_FILE")
    )

    # パスの検証
    if not result_output_folder:
        raise ValueError("result_output_folder が設定されていません")
    if not merged_csv_data_file:
        raise ValueError("merged_csv_data_file が設定されていません")

    logging.info(f"結果ディレクトリ: {result_output_folder}")
    logging.info(f"CSVファイル: {merged_csv_data_file}")

    try:
        # 処理対象の競技種目
        events = [
            (stroke, distance)
            for stroke in ["fly", "ba", "br", "fr", "im"]
            for distance in [50, 100, 200, 400]
        ]

        success_count = 0
        error_count = 0

        for stroke, distance in events:
            # Excelファイルのパス (IDが書き込まれたファイル)
            excel_file = os.path.join(
                result_output_folder, f"{distance}{stroke}_id.xlsx"
            )

            if not os.path.exists(excel_file):
                continue

            try:
                update_excel_with_player_data(
                    excel_file,
                    merged_csv_data_file,
                    result_output_folder,
                )
                logging.info(
                    f"✓ {stroke}{distance} のExcelファイルの更新が完了しました"
                )
                success_count += 1
                error_count += 1
            except Exception as e:
                logging.error(
                    f"✗ {stroke}{distance} の処理中に予期しないエラーが発生しました: {e}",
                    exc_info=True,
                )
                send_slack_message(
                    os.getenv("APP_NAME", "AquaProgrammer"),
                    f"{stroke}{distance} の処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}",
                )
                error_count += 1

        logging.info(f"処理完了 - 成功: {success_count}, エラー: {error_count}")

    except Exception as e:
        logging.error(
            f"メイン処理中に予期しないエラーが発生しました: {e}", exc_info=True
        )
        send_slack_message(
            os.getenv("APP_NAME", "AquaProgrammer"),
            f"メイン処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}",
        )
        raise


if __name__ == "__main__":
    main()
