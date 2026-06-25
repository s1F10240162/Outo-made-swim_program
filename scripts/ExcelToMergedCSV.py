import os
import pandas as pd
from typing import List, Optional, NoReturn
import logging
from dotenv import load_dotenv
import traceback
from module.send_message import send_slack_message

# ロガーの設定
logger = logging.getLogger(__name__)

def to_half_width(text):
    if not isinstance(text, str):
        if pd.isna(text):
            return ""
        return str(text)
    
    katakana_map = {
        "ガ": "ｶﾞ", "ギ": "ｷﾞ", "グ": "ｸﾞ", "ゲ": "ｹﾞ", "ゴ": "ｺﾞ",
        "ザ": "ｻﾞ", "ジ": "ｼﾞ", "ズ": "ｽﾞ", "ゼ": "ｾﾞ", "ゾ": "ｿﾞ",
        "ダ": "ﾀﾞ", "ヂ": "ﾁﾞ", "ヅ": "ﾂﾞ", "デ": "ﾃﾞ", "ド": "ﾄﾞ",
        "バ": "ﾊﾞ", "ビ": "ﾋﾞ", "ブ": "ﾌﾞ", "ベ": "ﾍﾞ", "ボ": "ﾎﾞ",
        "パ": "ﾊﾟ", "ピ": "ﾋﾟ", "プ": "ﾌﾟ", "ペ": "ﾍﾟ", "ポ": "ﾎﾟ",
        "ア": "ｱ", "イ": "ｲ", "ウ": "ｳ", "エ": "ｴ", "オ": "ｵ",
        "カ": "ｶ", "キ": "ｷ", "ク": "ｸ", "ケ": "ｹ", "コ": "ｺ",
        "サ": "ｻ", "シ": "ｼ", "ス": "ｽ", "セ": "ｾ", "ソ": "ｿ",
        "タ": "ﾀ", "チ": "ﾁ", "ツ": "ﾂ", "テ": "ﾃ", "ト": "ﾄ",
        "ナ": "ﾅ", "ニ": "ﾆ", "ヌ": "ﾇ", "ネ": "ﾈ", "ノ": "ﾉ",
        "ハ": "ﾊ", "ヒ": "ﾋ", "フ": "ﾌ", "ヘ": "ﾍ", "ホ": "ﾎ",
        "マ": "ﾏ", "ミ": "ﾐ", "ム": "ﾑ", "メ": "ﾒ", "モ": "ﾓ",
        "ヤ": "ﾔ", "ユ": "ﾕ", "ヨ": "ﾖ",
        "ラ": "ﾗ", "リ": "ﾘ", "ル": "ﾙ", "レ": "ﾚ", "ロ": "ﾛ",
        "ワ": "ﾜ", "ヲ": "ｦ", "ン": "ﾝ",
        "ァ": "ｧ", "ィ": "ｨ", "ゥ": "ｩ", "ェ": "ｪ", "ォ": "ｫ",
        "ッ": "ｯ", "ャ": "ｬ", "ュ": "ｭ", "ョ": "ｮ", "ヮ": "ﾜ",
        "ー": "ｰ", "ヴ": "ｳﾞ", "　": " "
    }
    
    res = text
    for f, h in katakana_map.items():
        res = res.replace(f, h)
        
    chars = []
    for c in res:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            chars.append(chr(code - 0xfee0))
        elif c == '　':
            chars.append(' ')
        else:
            chars.append(c)
    return "".join(chars).strip()

def delete_existing_csv(directory_path: str) -> List[str]:
    """
    指定フォルダ内のすべてのCSVファイルを削除する。
    
    引数:
        directory_path: CSVファイルを削除するディレクトリのパス
        
    戻り値:
        削除したファイル名のリスト
        
    例外:
        FileNotFoundError: ディレクトリが存在しない場合
        PermissionError: ファイルの削除権限がない場合
    """
    deleted_files = []
    try:
        files = os.listdir(directory_path)
        for filename in files:
            if filename.endswith(".csv"):
                file_path = os.path.join(directory_path, filename)
                try:
                    os.remove(file_path)
                    deleted_files.append(filename)
                    logger.info(f"ファイルを削除しました: {file_path}")
                except PermissionError as e:
                    logger.error(f"ファイルの削除権限がありません: {file_path} - {e}")
                    send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルの削除権限がありません: {file_path} - {e}")
        return deleted_files
    except FileNotFoundError as e:
        logger.error(f"ディレクトリが存在しません: {directory_path} - {e}")
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ディレクトリが存在しません: {directory_path} - {e}")
        return deleted_files
    except Exception as e:
        logger.error(f"delete_existing_csvで予期しないエラー: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"delete_existing_csvで予期しないエラー: {e}\n{traceback.format_exc()}")
        return deleted_files

def excel_to_csv(directory_path: str) -> List[str]:
    """
    フォルダ内のすべてのExcelファイルをCSVに変換する（既存ファイルは上書き）。
    
    引数:
        directory_path: Excelファイルを検索するディレクトリのパス
        
    戻り値:
        変換に成功したCSVファイル名のリスト
        
    例外:
        FileNotFoundError: ディレクトリまたはExcelファイルが存在しない場合
        ValueError: Excelファイルの読み込みに失敗した場合
    """
    import openpyxl
    converted_files = []
    try:
        files = os.listdir(directory_path)
        for filename in files:
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                if filename.startswith("~$"):
                    continue
                excel_file_path = os.path.join(directory_path, filename)
                try:
                    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
                    if "個人エントリー" not in wb.sheetnames:
                        logger.warning(f"個人エントリーシートが見つかりません: {excel_file_path}")
                        continue
                    
                    ws = wb["個人エントリー"]
                    
                    # 8, 9, 10行目のヘッダー情報を読み取る
                    row8 = [ws.cell(row=8, column=c).value for c in range(1, ws.max_column + 1)]
                    row9 = [ws.cell(row=9, column=c).value for c in range(1, ws.max_column + 1)]
                    row10 = [ws.cell(row=10, column=c).value for c in range(1, ws.max_column + 1)]
                    
                    # カラムヘッダーを再構築
                    headers = []
                    for col_idx in range(len(row8)):
                        r8 = str(row8[col_idx]).strip() if row8[col_idx] is not None else ""
                        r9 = str(row9[col_idx]).strip() if row9[col_idx] is not None else ""
                        r10 = str(row10[col_idx]).strip() if row10[col_idx] is not None else ""
                        
                        # 距離等の数値を整数へ正規化
                        if r9.endswith(".0"):
                            r9 = r9[:-2]
                        if r8.endswith(".0"):
                            r8 = r8[:-2]
                            
                        if col_idx == 0:
                            header = "No."
                        elif col_idx == 1:
                            header = "氏名"
                        elif col_idx == 2:
                            header = "ﾌﾘｶﾞﾅ"
                        elif col_idx == 3:
                            header = "学校名"
                        elif col_idx == 4:
                            header = "学年"
                        elif col_idx == 5:
                            header = "性別"
                        else:
                            # 競技列: パターンAとパターンBの両対応
                            if r8 and not r8.replace(".", "").isdigit() and r8 != "プログラムNo." and r8 != "None":
                                header = r8
                            elif r9 and r10 and r9 != "None" and r10 != "None":
                                header = f"{r9}{r10}"
                            else:
                                header = r8 if r8 else f"Col{col_idx+1}"
                        headers.append(header)
                        
                    # 12行目以降から選手データを抽出 (11行目の「例」はスキップ)
                    data_rows = []
                    for r in range(12, ws.max_row + 1):
                        name_val = ws.cell(row=r, column=2).value
                        if name_val is None or str(name_val).strip() == "":
                            continue
                        row_data = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                        data_rows.append(row_data)
                        
                    df = pd.DataFrame(data_rows, columns=headers)
                    if "ﾌﾘｶﾞﾅ" in df.columns:
                        df["ﾌﾘｶﾞﾅ"] = df["ﾌﾘｶﾞﾅ"].apply(to_half_width)
                    
                    csv_filename = f"{os.path.splitext(filename)[0]}_個人エントリー.csv"
                    csv_filepath = os.path.join(directory_path, csv_filename)
                    df.to_csv(csv_filepath, index=False, encoding="utf-8-sig")
                    converted_files.append(csv_filename)
                    logger.info(f"Excelファイルを変換しました: {excel_file_path} -> {csv_filepath}")
                    
                except Exception as e:
                    logger.error(f"Excelファイルの処理エラー: {excel_file_path} - {e}", exc_info=True)
                    send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excelファイルの処理エラー: {excel_file_path} - {e}\n{traceback.format_exc()}")
        return converted_files
    except FileNotFoundError as e:
        logger.error(f"ディレクトリが存在しません: {directory_path} - {e}")
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ディレクトリが存在しません: {directory_path} - {e}")
        return converted_files
    except Exception as e:
        logger.error(f"excel_to_csvで予期しないエラー: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"excel_to_csvで予期しないエラー: {e}\n{traceback.format_exc()}")
        return converted_files

def merge_csv_files(directory_path: str, output_file: str) -> Optional[pd.DataFrame]:
    """
    フォルダ内のCSVファイルを統合する。
    
    引数:
        directory_path: CSVファイルを検索するディレクトリのパス
        output_file: 結合したデータを保存する出力ファイルパス
        
    戻り値:
        結合したDataFrame、有効なデータがない場合はNone
        
    例外:
        FileNotFoundError: ディレクトリが存在しない場合
        PermissionError: 出力ファイルの書き込み権限がない場合
    """
    try:
        csv_files = [f for f in os.listdir(directory_path) if f.endswith(".csv")]
        concatenated_df = pd.DataFrame()
        processed_files = 0
        
        for file in csv_files:
            file_path = os.path.join(directory_path, file)
            try:
                df = pd.read_csv(file_path, encoding="utf-8-sig")
                if "氏名" not in df.columns:
                    logger.warning(f"CSVファイルに氏名列がありません: {file_path}")
                    continue
                df = df[df["氏名"].notna()]
                if not df.empty:
                    concatenated_df = pd.concat([concatenated_df, df], ignore_index=True)
                    logger.info(f"CSVファイルから有効なデータを抽出: {file_path}")
                    processed_files += 1
                else:
                    logger.info(f"CSVファイルに有効なデータがありません: {file_path}")
            except pd.errors.EmptyDataError as e:
                logger.error(f"CSVファイルが空です: {file_path} - {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"CSVファイルが空です: {file_path} - {e}")
            except Exception as e:
                logger.error(f"CSVファイル処理エラー: {file_path} - {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"CSVファイル処理エラー: {file_path} - {e}\n{traceback.format_exc()}")
        if not concatenated_df.empty:
            try:
                # 重複行を削除 (氏名, ﾌﾘｶﾞﾅ, 学校名, 学年, 性別 がすべて一致する行を重複とみなす)
                dedup_cols = ["氏名", "ﾌﾘｶﾞﾅ", "学校名", "学年", "性別"]
                existing_dedup_cols = [c for c in dedup_cols if c in concatenated_df.columns]
                if existing_dedup_cols:
                    before_count = len(concatenated_df)
                    concatenated_df = concatenated_df.drop_duplicates(subset=existing_dedup_cols, keep="first")
                    after_count = len(concatenated_df)
                    logger.info(f"重複データを削除しました: {before_count}行 -> {after_count}行")

                # IDを追加
                concatenated_df["ID"] = range(1, len(concatenated_df) + 1)
                # 出力ディレクトリが存在しない場合は作成
                os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
                concatenated_df.to_csv(output_file, index=False, encoding="utf-8-sig")
                logger.info(f"結合データを保存しました (処理ファイル数: {processed_files}): {output_file}")
                return concatenated_df
            except PermissionError as e:
                logger.error(f"出力ファイルの書き込み権限がありません: {output_file} - {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"出力ファイルの書き込み権限がありません: {output_file} - {e}")
                return None
            except Exception as e:
                logger.error(f"merge_csv_filesの保存処理で予期しないエラー: {e}", exc_info=True)
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"merge_csv_filesの保存処理で予期しないエラー: {e}\n{traceback.format_exc()}")
                return None
        else:
            logger.warning("結合可能な有効なデータがありませんでした")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), "結合可能な有効なデータがありませんでした")
            return None
            
    except FileNotFoundError as e:
        logger.error(f"ディレクトリが存在しません: {directory_path} - {e}")
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ディレクトリが存在しません: {directory_path} - {e}")
        return None
    except Exception as e:
        logger.error(f"merge_csv_filesで予期しないエラー: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"merge_csv_filesで予期しないエラー: {e}\n{traceback.format_exc()}")
        return None

def main(input_folder=None, output_csv=None) -> NoReturn:
    """
    メイン処理:
    1. フォルダ内の既存CSVを削除
    2. ExcelファイルをCSVに変換
    3. すべてのCSVを統合
    戻り値:
        なし
    """
    load_dotenv()
    input_folder = input_folder or os.getenv("INPUT_DATA_FILE")
    output_csv = output_csv or os.path.join(input_folder, os.getenv("MERGED_CSV_DATA_FILE"))
    try:
        if not input_folder:
            logger.error("環境変数 INPUT_DATA_FILE が設定されていません")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), "環境変数 INPUT_DATA_FILE が設定されていません")
            return
        os.makedirs(input_folder, exist_ok=True)
        delete_existing_csv(input_folder)
        excel_to_csv(input_folder)
        merge_csv_files(input_folder, output_csv)
        logger.info(f"すべての処理が完了しました: {input_folder}")
    except Exception as e:
        logger.error(f"ExcelToMergedCSVメイン処理で予期しないエラー: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ExcelToMergedCSVメイン処理で予期しないエラー: {e}\n{traceback.format_exc()}")

if __name__ == "__main__":
    main()
