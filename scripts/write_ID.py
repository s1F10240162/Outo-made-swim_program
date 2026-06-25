import os
import logging
import openpyxl
import glob
import re
import copy
from typing import List
from scripts.get_ID import get_player_id
from dotenv import load_dotenv
import traceback
from module.send_message import send_slack_message
from openpyxl.worksheet.pagebreak import Break


# ロガーの設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

def partition_names(names: List[str]) -> List[List[str]]:
    """
    名前リスト（早い順）から、最終組を必ず6名にするため、
    リストを反転（シート上で上が遅い、下が速い）し、
    余り（r = len % 6）があればそれを1組目（上側）に、
    残りを6名ずつのグループに分割する。
    ただし、1組目が1名だけになってしまう（r == 1 かつ全体の人数が7人以上）場合は、
    次の組から1名移動させて「1組目を2名、2組目を5名」とする。
    """
    # 入力は早い順なので反転して遅い順にする
    names_rev = names[::-1]
    total = len(names_rev)
    r = total % 6
    groups = []
    
    if r == 1 and total >= 7:
        # 1組目が1名になるのを防ぐため、1組目を2名、2組目を5名にする
        groups.append(names_rev[:2])
        groups.append(names_rev[2:7])
        start = 7
    else:
        start = 0
        if r != 0:
            groups.append(names_rev[:r])
            start = r
            
    while start < total:
        groups.append(names_rev[start : start + 6])
        start += 6
    return groups

def assign_group(group: List[str]) -> List[str]:
    """
    各グループ内で、中心から埋めるように名前を配置する。
    各グループは最終的に6コース（上からコース1～6）に対応するリストを返す。
    割り当てパターン（速い順に並べた場合のコース割り当て）は：
      最速 → コース4  
      2番目 → コース3  
      3番目 → コース5  
      4番目 → コース2  
      5番目 → コース6  
      6番目 → コース1  
    ※グループの人数が6未満の場合は、パターンの先頭分だけ割り当て、残りは空文字列とする。
    """
    pattern = [4, 3, 5, 2, 6, 1]  # 割り当てるコース番号（1～6）
    # グループ内を速い順（降順）にする
    sorted_group = group[::-1]
    # 結果はコース1～6の順（インデックス0～5）
    result = ["" for _ in range(6)]
    for i, name in enumerate(sorted_group):
        if i < len(pattern):
            course = pattern[i]
            result[course - 1] = name
    return result

def offset_formula(formula_str, row_offset):
    if not formula_str or not isinstance(formula_str, str) or not formula_str.startswith("="):
        return formula_str
    
    def replace_match_safe(m):
        col = m.group(1)
        row_str = m.group(2)
        
        # Check if there is $ before column
        col_start = m.start(1)
        if col_start > 0 and formula_str[col_start - 1] == '$':
            return m.group(0)
            
        # Check if there is $ before row
        row_start = m.start(2)
        if row_start > 0 and formula_str[row_start - 1] == '$':
            return m.group(0)
            
        # Check if followed by (
        end_idx = m.end()
        if end_idx < len(formula_str) and formula_str[end_idx] == '(':
            return m.group(0)
            
        row = int(row_str)
        return f"{col}{row + row_offset}"
        
    pattern = r'\b([A-Z]+)(\d+)\b'
    return re.sub(pattern, replace_match_safe, formula_str)

def write_to_excel(
    input_filename: str, 
    output_filename: str, 
    sheet_name: str, 
    cells: List[List[str]], 
    names: List[str],
    result_data_file: str
) -> bool:
    """
    指定したExcelのシートに選手IDを記入する。

    引数:
        - input_filename: 元となるExcelテンプレートファイル
        - output_filename: 出力するExcelファイル名
        - sheet_name: 書き込むシート名
        - cells: 書き込むセルのリスト（各リストはシート上のコース1～6に対応）
        - names: 書き込む選手IDリスト（早い順）

    戻り値:
        - 書き込みに成功した場合はTrue、失敗した場合はFalse
    """
    try:
        logger.info(f"Excelファイルを読み込み中: {input_filename}")
        try:
            wb = openpyxl.load_workbook(input_filename)
        except FileNotFoundError as e:
            logger.error(f"テンプレートファイルが見つかりません: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"テンプレートファイルが見つかりません: {input_filename} - {e}")
            return False
        except PermissionError as e:
            logger.error(f"テンプレートファイルへのアクセス権限がありません: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"テンプレートファイルへのアクセス権限がありません: {input_filename} - {e}")
            return False
        except Exception as e:
            logger.error(f"ファイルの読み込み中に予期しないエラーが発生しました: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルの読み込み中に予期しないエラーが発生しました: {input_filename} - {e}\n{traceback.format_exc()}")
            return False

        if sheet_name not in wb.sheetnames:
            logger.error(f"指定されたシート名が存在しません: '{sheet_name}'")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"指定されたシート名が存在しません: '{sheet_name}'")
            raise ValueError(f"シート名 '{sheet_name}' が見つかりません。")

        # 新しいワークブックを作成し、テンプレートシートの内容をコピーする
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 初期シート削除
        ws_original = wb[sheet_name]
        ws_new = new_wb.create_sheet(title=sheet_name)

        # 距離に応じたレイアウトパラメータを取得
        distance = int(re.search(r'\d+', sheet_name).group())
        is_short = distance in [50, 100]
        rows_per_page = 60 if is_short else 50
        heats_per_page = 12 if is_short else 5
        
        # 必要な列数のみコピーするように上限を設定
        if distance == 50:
            src_max_col = 14
            max_copy_col = 10
        elif distance == 100:
            src_max_col = 16
            max_copy_col = 10
        elif distance == 200:
            src_max_col = 11
            max_copy_col = 9
        else:  # 400m
            src_max_col = 15
            max_copy_col = 13

        # 列マッピングの定義
        col_map = {}
        if distance == 50:
            col_map = {
                2: 1,  # B (Lane Left) -> A
                4: 2,  # D (Name Left) -> B (ID written to B)
                5: 3,  # E (Furi Left) -> C
                6: 4,  # F (School Left) -> D
                7: 5,  # G (Time Left) -> E
                9: 6,  # I (Lane Right) -> F
                11: 7, # K (Name Right) -> G (ID written to G)
                12: 8, # L (Furi Right) -> H
                13: 9, # M (School Right) -> I
                14: 10 # N (Time Right) -> J
            }
        elif distance == 100:
            col_map = {
                2: 1,  # B (Lane Left) -> A
                4: 2,  # D (Name Left) -> B (ID written to B)
                5: 3,  # E (Furi Left) -> C
                6: 4,  # F (School Left) -> D
                8: 5,  # H (Time Left) -> E (ignore 7 G split)
                10: 6, # J (Lane Right) -> F (ignore 9 I spacer)
                12: 7, # L (Name Right) -> G (ID written to G)
                13: 8, # M (Furi Right) -> H
                14: 9, # School Right -> I
                16: 10 # P (Time Right) -> J (ignore 15 O split)
            }
        elif distance == 200:
            col_map = {
                2: 1,  # B (Lane) -> A
                4: 2,  # D (Name) -> B (ID written to B)
                5: 3,  # E (Furi) -> C
                6: 4,  # School -> D
                7: 5,  # Time -> E
                8: 6,  # Split 50m -> F
                9: 7,  # Split 100m -> G
                10: 8, # Split 150m -> H
                11: 9  # Split 200m -> I
            }
        else:  # 400m
            col_map = {
                2: 1,  # B (Lane) -> A
                4: 2,  # D (Name) -> B (ID written to B)
                5: 3,  # E (Furi) -> C
                6: 4,  # School -> D
                7: 5,  # Time -> E
                8: 6,  # Split 50m -> F
                9: 7,  # Split 100m -> G
                10: 8, # Split 150m -> H
                11: 9, # Split 200m -> I
                12: 10,# Split 250m -> J
                13: 11,# Split 300m -> K
                14: 12,# Split 350m -> L
                15: 13 # Split 400m -> M
            }

        # グループ分け（Heats）の件数を取得
        groups = partition_names(names)
        num_heats = len(groups)
        num_pages = max(1, (num_heats + heats_per_page - 1) // heats_per_page)

        # 列幅をコピー (指定された列数まで)
        for col_idx in range(1, max_copy_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            # この宛先列にマップされる元の列を探す
            src_cols = [src for src, dest in col_map.items() if dest == col_idx]
            if src_cols:
                src_col_idx = src_cols[0]
                src_col_letter = openpyxl.utils.get_column_letter(src_col_idx)
                width = ws_original.column_dimensions[src_col_letter].width
                if width:
                    ws_new.column_dimensions[col_letter].width = width
            else:
                ws_new.column_dimensions[col_letter].width = 8.43

        # Time列（E列とJ列）の幅を揃える（50m/100m用）
        if distance in [50, 100]:
            w5 = ws_new.column_dimensions['E'].width
            w10 = ws_new.column_dimensions['J'].width
            if w5 or w10:
                max_w = max(w5 or 0, w10 or 0, 8.43)
                ws_new.column_dimensions['E'].width = max_w
                ws_new.column_dimensions['J'].width = max_w

        # カナ列の幅を広げる (C列: 左組カナ, H列: 右組カナ)
        ws_new.column_dimensions['C'].width = 18.0
        if distance in [50, 100]:
            ws_new.column_dimensions['H'].width = 18.0

        # ページごとにテンプレートを複製してスタイルをコピー
        for p in range(num_pages):
            row_offset = p * rows_per_page
            
            # Determine how many rows to copy for this page
            if p < num_pages - 1:
                copy_rows = rows_per_page
            else:
                heats_on_page = num_heats - p * heats_per_page
                if is_short:
                    copy_rows = heats_on_page * 10 if heats_on_page <= 6 else 60
                else:
                    copy_rows = heats_on_page * 10
                copy_rows = max(10, copy_rows)  # 最低1ヒート分はコピーする
            
            # 各行データとスタイルをコピー
            for r in range(1, copy_rows + 1):
                src_height = ws_original.row_dimensions[r].height
                if src_height:
                    ws_new.row_dimensions[row_offset + r].height = src_height
                    
                for c in range(1, min(ws_original.max_column, src_max_col) + 1):
                    if c not in col_map:
                        continue
                    src_cell = ws_original.cell(row=r, column=c)
                    dest_c = col_map[c]
                    dest_cell = ws_new.cell(row=row_offset + r, column=dest_c)
                    
                    val = src_cell.value
                    if isinstance(val, str) and val.startswith("="):
                        dest_cell.value = offset_formula(val, row_offset)
                    else:
                        dest_cell.value = val
                        
                    if src_cell.has_style:
                        dest_cell.font = copy.copy(src_cell.font)
                        dest_cell.border = copy.copy(src_cell.border)
                        dest_cell.fill = copy.copy(src_cell.fill)
                        dest_cell.number_format = src_cell.number_format
                        dest_cell.alignment = copy.copy(src_cell.alignment)
            
            # 結合セルの情報をコピー
            for merged_range in ws_original.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if min_row <= copy_rows and max_row <= copy_rows:
                    if min_col in col_map and max_col in col_map:
                        new_min_row = min_row + row_offset
                        new_max_row = max_row + row_offset
                        new_min_col = col_map[min_col]
                        new_max_col = col_map[max_col]
                        ws_new.merge_cells(
                            start_row=new_min_row,
                            start_column=new_min_col,
                            end_row=new_max_row,
                            end_column=new_max_col
                        )

        # 改ページ（Page Breaks）を追加
        for p in range(num_pages - 1):
            break_row = (p + 1) * rows_per_page
            ws_new.row_breaks.append(Break(id=break_row))

        print("入力名前リスト（早い順）の最初の20件:", names[:20])
        print("グループごとの人数:", [len(g) for g in groups])
        # 各グループ内を、中心から埋める（各グループ結果はコース1～6の順）に再配置
        assigned_groups = [assign_group(group) for group in groups]
        print("各グループ再配置後（コース1～6順）の内容:")
        for idx, grp in enumerate(assigned_groups, start=1):
            print(f"組 {idx}: {grp}")

        # セル配置情報（cells）とグループごとの名前リストを対応付けて書き込み
        for cell_group, name_group in zip(cells, assigned_groups):
            for cell, name in zip(cell_group, name_group):
                ws_new[cell] = name


        try:
            os.makedirs(result_data_file, exist_ok=True)
            output_path = os.path.join(result_data_file, output_filename)
            new_wb.save(output_path)
            logger.info(f"Excelファイルを保存しました: {output_path}")
            return True
        except PermissionError as e:
            logger.error(f"ファイルへの書き込み権限がありません: {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルへの書き込み権限がありません: {e}")
            return False
        except Exception as e:
            logger.error(f"Excel書き込み処理中に予期しないエラーが発生しました: {e}", exc_info=True)
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excel書き込み処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
            return False
    except Exception as e:
        logger.error(f"Excel書き込み処理全体で予期しないエラーが発生しました: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excel書き込み処理全体で予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
        return False

def clean_output_directory(directory: str, pattern: str = "*.xlsx") -> int:
    """
    指定ディレクトリ内の特定パターンに一致するファイルを削除する。
    
    引数:
        - directory: 削除対象のディレクトリ
        - pattern: 削除対象のファイルパターン（デフォルトは *.xlsx）
        
    戻り値:
        - 削除したファイル数
    """
    deleted_count = 0
    try:
        if not os.path.exists(directory):
            logger.warning(f"削除対象ディレクトリが存在しません: {directory}")
            os.makedirs(directory, exist_ok=True)
            return 0

        file_pattern = os.path.join(directory, pattern)
        for file_path in glob.glob(file_pattern):
            try:
                os.remove(file_path)
                deleted_count += 1
                logger.debug(f"ファイルを削除しました: {file_path}")
            except PermissionError as e:
                logger.error(f"ファイルの削除権限がありません: {file_path} - {e}")
                print(f"警告: '{os.path.basename(file_path)}' の削除権限がありません")
            except Exception as e:
                logger.error(f"ファイル削除中にエラーが発生しました: {file_path} - {e}")

        logger.info(f"{directory} から {deleted_count}個のファイルを削除しました")
        return deleted_count
    except Exception as e:
        logger.error(f"ディレクトリのクリーン処理中にエラーが発生しました: {directory} - {e}")
        return deleted_count

def main(result_data_file=None, input_data_file=None, merged_csv_data_file=None, template_file=None) -> None:
    """
    メイン処理:
      1. 競技別に選手ID（名前）を取得
      2. Excelテンプレートの指定シートに選手IDを書き込む
      3. 各イベントごとに処理結果を出力する
    """
    # 引数優先、なければ環境変数
    load_dotenv()
    result_data_file = result_data_file or os.getenv("RESULT_DATA_FILE", "result")
    input_data_file = input_data_file or os.getenv("INPUT_DATA_FILE", "input_data_folder")
    merged_csv_data_file = merged_csv_data_file or os.path.join(input_data_file, os.getenv("MERGED_CSV_DATA_FILE", "merged_output.csv"))
    template_file = template_file or os.getenv("TEMPLATE_FILE", "template.xlsx")

    try:
        # 環境変数の検証
        if not result_data_file:
            logger.error("出力先パス(result_data_file)が指定されていません")
            print("エラー: 出力先パス(result_data_file)が指定されていません")
            return
        if not template_file:
            logger.error("テンプレートファイルが指定されていません")
            print("エラー: テンプレートファイルが指定されていません")
            return
        if not os.path.exists(template_file):
            logger.error(f"テンプレートファイルが見つかりません: {template_file}")
            print(f"エラー: テンプレートファイル '{template_file}' が見つかりません")
            return
        if not merged_csv_data_file or not os.path.exists(merged_csv_data_file):
            logger.error(f"マージされたCSVファイルが見つかりません: {merged_csv_data_file}")
            print(f"エラー: マージされたCSVファイル '{merged_csv_data_file}' が見つかりません")
            return

        # 出力フォルダを確認・作成
        if not os.path.exists(result_data_file):
            os.makedirs(result_data_file, exist_ok=True)
            logger.info(f"出力フォルダを作成しました: {result_data_file}")
            print(f"出力フォルダを作成しました: {result_data_file}")

        # 出力フォルダ内の過去データ（Excelファイル）を削除
        deleted_count = clean_output_directory(result_data_file)
        print(f"{result_data_file}内の過去データである.xlsxファイルを{deleted_count}件削除しました。")

        # 各距離に対するセル設定（必要な距離のみ設定）
        cell_config = {
            50: ["C", "J", "Q", "X", "AE", "AL", "AS", "AZ", "BG", "BN"],
            100: ["C", "K", "R", "Z", "AG", "AO", "AU", "BI", "BP"],
            200: ["C", "M", "W", "AG", "AQ"],
            400: ["C", "Q", "AE"],
        }

        # 全種目の組み合わせを生成
        events = [
            (stroke, distance)
            for stroke in ["im", "fly", "ba", "br", "fr"]
            for distance in [50, 100, 200, 400]
            if distance in cell_config  # セル設定がある距離のみ対象
        ]

        successful_events = 0
        failed_events = 0

        # 各種目ごとに選手IDを取得し、Excelに書き込む
        for stroke, distance in events:
            logger.info(f"イベント {stroke}{distance} の処理を開始")

            # 選手ID（ここでは名前リスト）を取得する
            ids = get_player_id(merged_csv_data_file, (stroke, distance), category="mixed")
            if not ids:
                logger.warning(f"イベント {stroke}{distance} のIDデータが見つかりません")
                # send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} のIDデータが見つかりません。")
                failed_events += 1
                continue

            # グループ分け（Heats）の件数を取得
            groups = partition_names(ids)
            num_heats = len(groups)
            
            # 各ヒート（組）に対するセル座標を計算
            cells = []
            for h_idx in range(num_heats):
                if distance in [50, 100]:
                    page = h_idx // 12
                    in_page_idx = h_idx % 12
                    p_heat = in_page_idx % 6
                    is_right = in_page_idx >= 6
                    
                    col_letter = "G" if is_right else "B"
                    
                    start_row = page * 60 + 4 + p_heat * 10
                else:  # 200m, 400m
                    page = h_idx // 5
                    p_heat = h_idx % 5
                    col_letter = "B"
                    
                    start_row = page * 50 + 4 + p_heat * 10
                    
                heat_cells = [f"{col_letter}{start_row + j}" for j in range(6)]
                cells.append(heat_cells)

            output_filename = f"{distance}{stroke}_id.xlsx"
            sheet_name = f"{distance}m"
            try:
                success = write_to_excel(template_file, output_filename, sheet_name, cells, ids, result_data_file)
                if success:
                    successful_events += 1
                    logger.info(f"イベント {stroke}{distance} の処理が成功しました")
                else:
                    failed_events += 1
                    logger.error(f"イベント {stroke}{distance} の処理が失敗しました")
            except FileNotFoundError as e:
                logger.error(f"ファイルが見つかりません: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルが見つかりません: {e}")
                failed_events += 1
                continue
            except ValueError as e:
                logger.error(f"イベント {stroke}{distance} の処理中に値エラーが発生しました: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} の処理中に値エラーが発生しました: {e}")
                failed_events += 1
                continue
            except Exception as e:
                logger.error(f"イベント {stroke}{distance} の処理中に予期しないエラーが発生しました: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} の処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
                failed_events += 1
                continue

        # 処理結果のサマリーを出力
        logger.info(f"処理完了: 成功={successful_events}件, 失敗={failed_events}件")
        if successful_events > 0:
            logger.info(f"処理が完了しました。{successful_events}件のイベントを正常に処理しました。")
        if failed_events > 0:
            logger.warning(f"警告: {failed_events}件のイベントで処理に失敗しました。")

    except Exception as e:
        logger.error(f"予期しないエラーが発生しました: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")

if __name__ == "__main__":
    main()
