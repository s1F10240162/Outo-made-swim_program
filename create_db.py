import pandas as pd
import glob
import os
from openpyxl import load_workbook, Workbook

# ===== 設定 =====
input_xlsx_folder = "input_data_folder/"
template_path = "template.xlsx"
output_csv_path = "output_data/entries.csv"
output_templates = "output_data/output_templates/"

os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
os.makedirs(output_templates, exist_ok=True)

# ===== STEP1: Excelファイルを読み込み、統合DataFrameを作成 =====
xlsx_paths = glob.glob(os.path.join(input_xlsx_folder, "*.xlsx"))
column_sets = []
for path in xlsx_paths:
    df_raw = pd.ExcelFile(path).parse("個人エントリー", header=None)
    header_row = df_raw.iloc[7]
    column_sets.append(set(header_row.dropna()))

base_cols = ['氏名', 'ﾌﾘｶﾞﾅ', '学校名', '学年', '性別']
all_columns = set.union(*column_sets)
event_cols = sorted(str(c) for c in (all_columns - set(base_cols) - {'種目数'}) if isinstance(c, str))
final_columns = ['id', 'member_num'] + base_cols + event_cols + ['種目数']
all_df = pd.DataFrame(columns=final_columns)

for path in xlsx_paths:
    try:
        raw = pd.ExcelFile(path).parse("個人エントリー", header=None)
        hdr = raw.iloc[7]
        df = pd.DataFrame(raw.iloc[10:].values, columns=hdr)

        df = df[df['氏名'] != '駒場　太郎']
        df = df[df['氏名'].notna()]
        if df['ﾌﾘｶﾞﾅ'].isna().any():
            raise ValueError("フリガナ欠損があります")

        shomoku = df.pop('種目数') if '種目数' in df.columns else None
        df.insert(0, 'member_num', range(1, len(df)+1))
        df.insert(0, 'id', range(len(all_df)+1, len(all_df)+1+len(df)))
        for col in final_columns:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[final_columns]
        if shomoku is not None:
            df['種目数'] = shomoku.values

        all_df = pd.concat([all_df, df], ignore_index=True)

    except Exception as e:
        print(f"⚠ {os.path.basename(path)} をスキップ: {e}")

all_df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
print(f"✔ 統合CSVを保存: {output_csv_path}")

# ===== STEP2: 種目別テンプレート書き出し（6コース、センターアウト） =====
for ev in event_cols:
    try:
        part = all_df[all_df[ev].notna()].copy()
        if len(part) > 175:
            raise ValueError(f"{ev} の出場者が175人超過")

        distance = ''.join(filter(str.isdigit, ev))
        stroke = ''.join(filter(str.isalpha, ev))

        def convert_to_seconds(time_val):
            try:
                time_val = float(time_val)
                minutes = int(time_val // 100)
                seconds = time_val - (minutes * 100)
                return minutes * 60 + seconds
            except:
                return float('inf')

        part['タイム秒'] = part[ev].apply(convert_to_seconds)
        part = part.sort_values(by='タイム秒', ascending=False).reset_index(drop=True)

        wb_template = load_workbook(template_path)
        sheet = f"{distance}m"
        if sheet not in wb_template.sheetnames:
            raise ValueError(f"{sheet} シートが存在しません")

        new_wb = Workbook()
        new_ws = wb_template[sheet]
        new_sheet = new_wb.active
        new_sheet.title = sheet

        for row in new_ws.iter_rows():
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value

        spacing_adjust = {
            "50": 0,
            "100": 1,
            "200": 3,
            "400": 7
        }
        col_spacing = spacing_adjust.get(distance, 0)

        total_participants = len(part)
        lanes_per_heat = 6
        full_groups = total_participants // lanes_per_heat
        remainder = total_participants % lanes_per_heat
        total_groups = full_groups + (1 if remainder > 0 else 0)

        group_assignments = []
        idx = 0
        if remainder > 0:
            group_assignments.append(part.iloc[idx:idx+remainder])
            idx += remainder
        for _ in range(full_groups):
            group_assignments.append(part.iloc[idx:idx+lanes_per_heat])
            idx += lanes_per_heat

        center_out_order = [3, 2, 4, 1, 5, 0]

        for group_index, group_df in enumerate(reversed(group_assignments)):
            actual_group = total_groups - 1 - group_index
            group_df = group_df.reset_index(drop=True)
            for logical_lane_index, true_lane_index in enumerate(center_out_order[:len(group_df)]):
                row = group_df.iloc[logical_lane_index]
                row_blk = (actual_group % 6) * 10
                col_blk = (actual_group // 6) * (7 + col_spacing)
                r = row_blk + true_lane_index + 4
                c = col_blk + 2
                new_sheet.cell(row=r, column=c, value=row['氏名'])
                new_sheet.cell(row=r, column=c+1, value=row['ﾌﾘｶﾞﾅ'])
                new_sheet.cell(row=r, column=c+2, value=row['学校名'])
                new_sheet.cell(row=r, column=c+3, value=row['学年'])

        out_xlsx = os.path.join(output_templates, f"{stroke}_{distance}.xlsx")
        new_wb.save(out_xlsx)
        print(f"✔ 出力完了: {out_xlsx}")

    except Exception as e:
        print(f"⚠ {ev} 処理エラー: {e}")
